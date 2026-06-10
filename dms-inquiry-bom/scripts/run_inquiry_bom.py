#!/usr/bin/env python3
"""BOM清单Excel生成脚本。

支持单个或批量生成BOM清单文件，格式匹配产品物料导入模板。

用法：
  # 单个BOM
  python run_inquiry_bom.py --name "覃建发" --components 30 --items '[["6B001492",30],["AA001653",1]]'

  # 批量BOM
  python run_inquiry_bom.py --bom-list '[
    {"name": "覃建发", "components": 30, "items": [["6B001492",30],["AA001653",1]]},
    {"name": "蔡敏捷", "components": 185, "items": [["6B001492",185],["AB001347",2],["AB001067",1]]}
  ]' --output-dir "."
"""

import argparse
import json
import os
import sys
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

# 修复 Windows 中文乱码
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import _compat  # noqa: F401, E402


def generate_bom(name: str, components: int, items: list, output_dir: str = ".",
                 project: str = None) -> str:
    """生成单个BOM清单Excel文件。

    Args:
        name: 业务员姓名
        components: 组件总块数
        items: [[物料编号, 数量], ...]
        output_dir: 输出目录
        project: 项目名称（可选，用于文件命名）

    Returns:
        生成的文件路径
    """
    today = date.today().strftime("%Y%m%d")
    # 截取项目名称的关键部分（去掉"分布式光伏发电项目"等后缀）
    if project:
        # 生成文件名的项目名简写（去掉已知的后缀关键词）
        project_short = project
        for keyword in ['分布式光伏发电项目', '光伏发电项目', '分布式光伏', '光伏项目', '发电项目']:
            idx = project_short.find(keyword)
            if idx >= 0:
                project_short = project_short[:idx]
                break
        project_short = project_short[:15].rstrip()
        filename = f"{name}{components}块组件{project_short}{today}.xlsx"
    else:
        filename = f"{name}{components}块组件{today}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws["A1"] = "物料编号"
    ws["B1"] = "数量"
    ws["A1"].font = header_font
    ws["B1"].font = header_font
    ws["A1"].border = thin_border
    ws["B1"].border = thin_border
    ws["A1"].alignment = header_align
    ws["B1"].alignment = header_align

    for i, (material_id, qty) in enumerate(items, start=2):
        ws[f"A{i}"] = material_id
        ws[f"B{i}"] = qty
        ws[f"A{i}"].border = thin_border
        ws[f"B{i}"].border = thin_border
        ws[f"B{i}"].alignment = header_align

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 10

    wb.save(filepath)
    return filepath


def generate_multiple(bom_list: list, output_dir: str = ".") -> list:
    """批量生成BOM清单Excel文件。

    Args:
        bom_list: [{"name": "...", "components": N, "items": [[...], ...], "project": "..."}, ...]
        output_dir: 输出目录

    Returns:
        生成的文件路径列表
    """
    paths = []
    for spec in bom_list:
        name = spec["name"]
        components = spec["components"]
        items = spec["items"]
        project = spec.get("project")  # 可选的项目名称
        path = generate_bom(name, components, items, output_dir, project)
        paths.append(path)
        print(f"  {path}", file=sys.stderr)
    return paths


def parse_items(items_str: str) -> list:
    """解析items参数，支持两种格式：
    1. JSON数组格式: [["6B001492",30],["AA001653",1]]
    2. 简洁格式: 6B001492:30,AA001653:1

    Returns:
        [[物料编号, 数量], ...]
    """
    items_str = items_str.strip()

    # 尝试JSON格式
    if items_str.startswith('['):
        return json.loads(items_str)

    # 简洁格式: code:qty,code:qty
    items = []
    for pair in items_str.split(','):
        pair = pair.strip()
        if ':' in pair:
            code, qty = pair.split(':', 1)
            items.append([code.strip(), int(qty.strip())])
        elif 'x' in pair.lower():
            code, qty = pair.lower().split('x', 1)
            items.append([code.strip(), int(qty.strip())])

    return items


def main():
    parser = argparse.ArgumentParser(description="生成BOM清单Excel")
    parser.add_argument("--name", help="业务员姓名（单个模式）")
    parser.add_argument("--components", type=int, help="组件总块数（单个模式）")
    parser.add_argument("--items", help='物料清单，支持两种格式：\n'
                                         '  JSON: [["6B001492",30],["AA001653",1]]\n'
                                         '  简洁: 6B001492:30,AA001653:1')
    parser.add_argument("--project", help="项目名称（可选，用于文件命名）")
    parser.add_argument("--bom-list", help='批量模式：JSON数组，每项含 name/components/items/project')
    parser.add_argument("--output-dir", default=".", help="输出目录")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    if args.bom_list:
        # 批量模式
        bom_list = json.loads(args.bom_list)
        print(f"[BOM] 批量生成 {len(bom_list)} 个文件:", file=sys.stderr)
        paths = generate_multiple(bom_list, args.output_dir)
        # stdout输出文件路径列表（供Claude读取）
        print(json.dumps(paths, ensure_ascii=False))
    elif args.name and args.components is not None and args.items:
        # 单个模式
        items = parse_items(args.items)
        path = generate_bom(args.name, args.components, items, args.output_dir, args.project)
        print(f"[BOM] 已生成: {path}", file=sys.stderr)
        print(path)
    else:
        print("[错误] 请提供 --name/--components/--items（单个模式）或 --bom-list（批量模式）", file=sys.stderr)
        raise SystemExit(1)


if __name__ == "__main__":
    main()
