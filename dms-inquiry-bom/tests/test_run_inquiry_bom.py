#!/usr/bin/env python3
"""generate_bom.py 的完整测试套件。

覆盖范围：
  1. generate_bom 的基本功能
  2. generate_multiple 批量模式
  3. parse_items 参数解析（JSON 和简洁格式）
  4. 项目名截断规则
  5. 输出文件路径格式
"""

import json
import os
import sys
import tempfile
import unittest

TEST_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(TEST_DIR)
SCRIPTS_DIR = os.path.join(SKILL_DIR, "scripts")
sys.path.insert(0, SCRIPTS_DIR)

import _compat  # noqa: F401, E402

from generate_bom import (  # noqa: E402
    generate_bom,
    generate_multiple,
    parse_items,
)


# ═══════════════════════════════════════════════════════════
#  1. parse_items 参数解析
# ═══════════════════════════════════════════════════════════

class TestParseItems(unittest.TestCase):
    """测试 parse_items 函数"""

    def test_json_array_format(self):
        """JSON 数组格式: [["6B001492",30],["AA001653",1]]"""
        result = parse_items('[["6B001492",30],["AA001653",1]]')
        self.assertEqual(result, [["6B001492", 30], ["AA001653", 1]])

    def test_simple_format(self):
        """简洁格式: 6B001492:30,AA001653:1"""
        result = parse_items("6B001492:30,AA001653:1")
        self.assertEqual(result, [["6B001492", 30], ["AA001653", 1]])

    def test_x_format(self):
        """简洁格式（x分隔）: 6B001492x30,AA001653x1
        注意：parse_items 对 x 格式会做 .lower() 处理"""
        result = parse_items("6B001492x30,AA001653x1")
        # lower() 处理后物料编码会变成小写
        self.assertEqual(result, [["6b001492", 30], ["aa001653", 1]])

    def test_single_item(self):
        """单物料"""
        result = parse_items('[["6B001492",30]]')
        self.assertEqual(result, [["6B001492", 30]])

    def test_empty_json_array(self):
        """空JSON数组"""
        result = parse_items("[]")
        self.assertEqual(result, [])


# ═══════════════════════════════════════════════════════════
#  2. generate_bom 基本功能
# ═══════════════════════════════════════════════════════════

class TestGenerateBom(unittest.TestCase):
    """测试 generate_bom"""

    def setUp(self):
        self.tmp_dir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_bom_generated(self):
        """BOM 文件成功生成"""
        path = generate_bom("张三", 800, [["6B001492", 800], ["AB001347", 8]],
                            output_dir=self.tmp_dir)
        self.assertTrue(os.path.exists(path))
        self.assertIn("张三", path)
        self.assertIn("800块组件", path)

    def test_bom_without_project(self):
        """无项目名时文件名不含项目名"""
        path = generate_bom("张三", 800, [["6B001492", 800]],
                            output_dir=self.tmp_dir)
        # 文件名格式：张三800块组件20260610.xlsx，不含项目名
        self.assertIn("张三800块组件", path)
        # 路径中不应出现"项目"字样（"组件"不算）
        basename = os.path.basename(path)
        if "项目" in basename:
            self.assertFalse(basename.index("项目") < basename.index("组"),
                             "文件名中 '项目' 应在 '组' 之后（即属于 '组件'）")

    def test_bom_with_project(self):
        """含项目名时文件名包含项目简称"""
        path = generate_bom("张三", 800, [["6B001492", 800]],
                            project="广西南宁江南区分布式光伏发电项目",
                            output_dir=self.tmp_dir)
        self.assertIn("广西南宁江南区", path)
        # 不应包含"分布式光伏发电项目"
        self.assertNotIn("分布式光伏发电项目", path)

    def test_bom_project_name_not_stripping_locality(self):
        """项目名不应被错误截断地名中的'区'字"""
        path = generate_bom("张三", 800, [["6B001492", 800]],
                            project="南宁市兴宁区分布式光伏发电项目",
                            output_dir=self.tmp_dir)
        # "兴宁区"的"区"字应保留
        self.assertIn("南宁市兴宁区", path,
                       "项目名中的'区'不应被 rstrip 掉")

    def test_bom_project_long_name(self):
        """长项目名被截断到15字符"""
        long_name = "这是一个非常长的项目名称测试截断功能"
        path = generate_bom("张三", 800, [["6B001492", 800]],
                            project=long_name,
                            output_dir=self.tmp_dir)
        # 文件名应在15个中文字符后截断
        self.assertIn("这是一个非常长的", path)
        # 不应包含截断后的完整名称
        self.assertNotIn("截断功能", path)

    def test_bom_items_in_excel(self):
        """Excel 中正确写入物料数据"""
        import openpyxl
        path = generate_bom("张三", 800, [["6B001492", 800], ["AB001347", 8]],
                            output_dir=self.tmp_dir)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        self.assertEqual(ws["A1"].value, "物料编号")
        self.assertEqual(ws["B1"].value, "数量")
        self.assertEqual(ws["A2"].value, "6B001492")
        self.assertEqual(ws["B2"].value, 800)
        self.assertEqual(ws["A3"].value, "AB001347")
        self.assertEqual(ws["B3"].value, 8)


# ═══════════════════════════════════════════════════════════
#  3. generate_multiple 批量模式
# ═══════════════════════════════════════════════════════════

class TestGenerateMultiple(unittest.TestCase):
    """测试 generate_multiple"""

    def setUp(self):
        self.tmp_dir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_batch_generation(self):
        """批量生成多个BOM"""
        bom_list = [
            {"name": "张三", "components": 800,
             "items": [["6B001492", 800]]},
            {"name": "李四", "components": 300,
             "items": [["6B001492", 300]]},
        ]
        paths = generate_multiple(bom_list, output_dir=self.tmp_dir)
        self.assertEqual(len(paths), 2)
        for p in paths:
            self.assertTrue(os.path.exists(p))

    def test_batch_with_project(self):
        """批量模式支持 project 字段"""
        bom_list = [
            {"name": "张三", "components": 800,
             "items": [["6B001492", 800]],
             "project": "南宁分布式光伏项目"},
        ]
        paths = generate_multiple(bom_list, output_dir=self.tmp_dir)
        self.assertEqual(len(paths), 1)
        self.assertIn("南宁", paths[0])


# ═══════════════════════════════════════════════════════════
#  入口
# ═══════════════════════════════════════════════════════════

if __name__ == "__main__":
    unittest.main()
