"""Excel 样式主题配置 — 靛蓝现代报表风格（匹配 HTML 报表配色）。

设计来源：references/report_template.html CSS Design Tokens
配色体系：靛蓝主色 (#1D4ED8) + 翠绿/琥珀/朱红功能色 + 石板暖灰中性色
"""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ==================== 通用边框 ====================

THIN_BORDER = Border(
    left=Side("thin", color="E2E8F0"),
    right=Side("thin", color="E2E8F0"),
    top=Side("thin", color="E2E8F0"),
    bottom=Side("thin", color="E2E8F0"),
)

HEADER_BORDER = Border(
    left=Side("thin", color="1D4ED8"),
    right=Side("thin", color="1D4ED8"),
    top=Side("thin", color="1D4ED8"),
    bottom=Side("medium", color="1E3A8A"),
)

BOTTOM_BORDER = Border(
    bottom=Side("medium", color="1D4ED8"),
    left=Side("thin", color="E2E8F0"),
    right=Side("thin", color="E2E8F0"),
    top=Side("thin", color="E2E8F0"),
)

CARD_BORDER = Border(
    left=Side("thin", color="CBD5E1"),
    right=Side("thin", color="CBD5E1"),
    top=Side("thin", color="CBD5E1"),
    bottom=Side("thin", color="CBD5E1"),
)

# ==================== 调色板 ====================

class Colors:
    # 主色系 — 靛蓝（匹配 HTML --primary）
    DARK_BLUE = "1E3A8A"        # 靛蓝-深色
    PRIMARY_BLUE = "1D4ED8"     # 靛蓝-主色
    ACCENT_BLUE = "3B82F6"      # 靛蓝-亮色
    PRIMARY_LIGHT = "60A5FA"    # 靛蓝-浅色
    PRIMARY_L10 = "DBEAFE"      # 10% 浅色 tint
    PRIMARY_L5 = "EFF6FF"       # 5% 浅色 tint

    # 功能色（匹配 HTML）
    GREEN = "10B981"            # 翠绿
    ORANGE = "F59E0B"           # 琥珀
    RED = "EF4444"              # 朱红
    PURPLE = "8B5CF6"           # 紫罗兰
    TEAL = "06B6D4"             # 青碧

    # 中性色 — 石板暖灰系（匹配 HTML）
    TEXT_PRIMARY = "0F172A"     # 石板-900
    TEXT_SECONDARY = "475569"   # 石板-600
    TEXT_TERTIARY = "94A3B8"    # 石板-400
    BG_PAGE = "F1F5F9"          # 石板-100
    BG_CARD = "F8FAFC"          # 石板-50
    BORDER = "CBD5E1"           # 石板-300
    BORDER_LIGHT = "E2E8F0"     # 石板-200
    WHITE = "FFFFFF"

    # 表格交替行（匹配 HTML tbody tr:nth-child(even)）
    ROW_ALT = "F7F8FA"

# ==================== 字体 ====================

FONT_TITLE = Font(bold=True, size=14, color=Colors.PRIMARY_BLUE, name="微软雅黑")
FONT_SECTION = Font(bold=True, size=11, color=Colors.PRIMARY_BLUE, name="微软雅黑")
FONT_SUBSECTION = Font(bold=True, size=11, color=Colors.PRIMARY_BLUE, name="微软雅黑")
FONT_HEADER = Font(bold=True, size=10, color=Colors.WHITE, name="微软雅黑")
FONT_DATA = Font(size=10, color=Colors.TEXT_PRIMARY, name="微软雅黑")
FONT_LABEL = Font(bold=False, size=10, color=Colors.TEXT_PRIMARY, name="微软雅黑")
FONT_KPI_BIG = Font(bold=True, size=10, color=Colors.PRIMARY_BLUE, name="微软雅黑")
FONT_KPI_MED = Font(bold=True, size=10, color=Colors.PRIMARY_BLUE, name="微软雅黑")
FONT_HINT = Font(size=9, color=Colors.TEXT_TERTIARY, name="微软雅黑")
FONT_VALUE = Font(bold=True, size=10, color=Colors.PRIMARY_BLUE, name="微软雅黑")

# 功能色字体 — KPI 数值用颜色区分
FONT_GREEN = Font(bold=True, size=10, color=Colors.GREEN, name="微软雅黑")
FONT_ORANGE = Font(bold=True, size=10, color=Colors.ORANGE, name="微软雅黑")
FONT_RED = Font(bold=True, size=10, color=Colors.RED, name="微软雅黑")

# ==================== 填充 ====================

FILL_HEADER = PatternFill(start_color=Colors.PRIMARY_BLUE, end_color=Colors.PRIMARY_BLUE, fill_type="solid")
FILL_LIGHT = PatternFill(start_color=Colors.PRIMARY_L10, end_color=Colors.PRIMARY_L10, fill_type="solid")
FILL_VERY_LIGHT = PatternFill(start_color=Colors.ROW_ALT, end_color=Colors.ROW_ALT, fill_type="solid")
FILL_CARD = PatternFill(start_color=Colors.BG_CARD, end_color=Colors.BG_CARD, fill_type="solid")
FILL_WHITE = PatternFill(start_color=Colors.WHITE, end_color=Colors.WHITE, fill_type="solid")

# ==================== 对齐 ====================

ALIGN_CENTER = Alignment(vertical="center", horizontal="center", wrap_text=False)
ALIGN_LEFT = Alignment(vertical="center", horizontal="left", wrap_text=False)
ALIGN_RIGHT = Alignment(vertical="center", horizontal="right", wrap_text=False)
ALIGN_HEADER = Alignment(wrap_text=True, vertical="center", horizontal="center")
ALIGN_DATA = Alignment(vertical="center", horizontal="center", wrap_text=False)
ALIGN_DATA_LEFT = Alignment(vertical="center", horizontal="left", wrap_text=True)

# ==================== 行高 / 列宽 ====================

ROW_HEIGHT_TITLE = 42
ROW_HEIGHT_SECTION = 32
ROW_HEIGHT_DATA = 28
ROW_HEIGHT_HEADER = 36

# 询价汇总 Sheet 列宽
COLUMN_WIDTHS: list[int] = [
    22, 50, 16, 45, 16, 20, 16, 18, 18,
    14, 16, 24, 28, 12, 14, 14, 14, 14, 24,
]


def apply_header_style(ws: Any, row: int, headers: list[str], start_col: int = 1) -> None:
    """给指定行应用深蓝表头样式。"""
    for col, h in enumerate(headers, start_col):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.border = HEADER_BORDER
        cell.alignment = ALIGN_HEADER


def apply_data_row(ws: Any, row: int, values: list, start_col: int = 1, is_alt: bool = False) -> None:
    """给数据行应用标准样式，无背景色，全部居中。"""
    for col, val in enumerate(values, start_col):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = FONT_DATA
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER
        cell.fill = FILL_WHITE


def write_section_title(ws: Any, row: int, col: int, title: str, merge_end_col: int | None = None) -> int:
    """写入节标题，蓝色文字 + 底部线条，可选合并单元格，返回下一行号。"""
    font = Font(bold=True, size=11, color=Colors.PRIMARY_BLUE, name="微软雅黑")

    cell = ws.cell(row=row, column=col, value=title)
    cell.font = font
    cell.fill = FILL_WHITE
    cell.alignment = ALIGN_LEFT
    cell.border = Border(bottom=Side("medium", color=Colors.PRIMARY_L10))

    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
        for c in range(col, merge_end_col + 1):
            ws.cell(row=row, column=c).fill = FILL_WHITE
            ws.cell(row=row, column=c).font = font

    ws.row_dimensions[row].height = ROW_HEIGHT_SECTION - 8
    return row + 1


def write_kpi_card(ws: Any, row: int, col: int, label: str, value: str | int | float,
                   unit: str = "", value_font: Font | None = None,
                   alt_fill: bool = False) -> None:
    """写入 KPI 卡片，所有内容居中对齐。"""
    fill = FILL_WHITE
    border = CARD_BORDER

    # label
    c1 = ws.cell(row=row, column=col, value=label)
    c1.font = FONT_LABEL
    c1.alignment = ALIGN_CENTER
    c1.border = border
    c1.fill = fill

    # value
    c2 = ws.cell(row=row, column=col + 1, value=value)
    c2.font = value_font or FONT_KPI_BIG
    c2.alignment = ALIGN_CENTER
    c2.border = border
    c2.fill = fill

    # unit
    if unit:
        c3 = ws.cell(row=row, column=col + 2, value=unit)
        c3.font = FONT_HINT
        c3.alignment = ALIGN_CENTER
        c3.border = border
        c3.fill = fill

    ws.row_dimensions[row].height = ROW_HEIGHT_DATA + 4


def apply_accent_row(ws: Any, row: int, values: list, start_col: int = 1, color="ACCENT_BLUE") -> None:
    """给重要数据行应用强调样式。"""
    accent_color = getattr(Colors, color, Colors.ACCENT_BLUE)
    for col, val in enumerate(values, start_col):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True, size=11, color=accent_color, name="微软雅黑")
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER
        cell.fill = FILL_LIGHT

    ws.row_dimensions[row].height = ROW_HEIGHT_DATA


def apply_status_cell(ws: Any, row: int, col: int, status: str, bg_color: str) -> None:
    """写入状态指示器单元格。"""
    cell = ws.cell(row=row, column=col, value=status)
    cell.font = Font(bold=True, size=10, color=Colors.WHITE, name="微软雅黑")
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.border = CARD_BORDER
    cell.alignment = ALIGN_CENTER
