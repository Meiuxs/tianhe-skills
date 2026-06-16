"""excel_styles.py 单元测试。"""

import sys
from pathlib import Path
import unittest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).parent.parent))
from excel_styles import (
    Colors,
    FONT_TITLE, FONT_SECTION, FONT_HEADER, FONT_DATA,
    FONT_KPI_BIG, FONT_KPI_MED, FONT_HINT,
    FILL_HEADER, FILL_LIGHT, FILL_VERY_LIGHT, FILL_CARD,
    ALIGN_CENTER, ALIGN_LEFT, ALIGN_HEADER, ALIGN_DATA,
    ROW_HEIGHT_TITLE, ROW_HEIGHT_SECTION, ROW_HEIGHT_DATA, ROW_HEIGHT_HEADER,
    THIN_BORDER, HEADER_BORDER, BOTTOM_BORDER, CARD_BORDER,
    apply_header_style, apply_data_row,
    write_section_title, write_kpi_card,
)

class TestColorsClass(unittest.TestCase):
    def test_color_values_are_hex(self):
        for attr_name in dir(Colors):
            if not attr_name.startswith("_"):
                value = getattr(Colors, attr_name)
                if isinstance(value, str):
                    self.assertRegex(value, r"^[0-9A-F]{6}$", f"{attr_name} not hex")
    def test_key_colors_exist(self):
        for c in ["DARK_BLUE", "PRIMARY_BLUE", "ACCENT_BLUE", "WHITE", "TEXT_PRIMARY"]:
            self.assertTrue(hasattr(Colors, c))

class TestFontDefinitions(unittest.TestCase):
    def test_fonts_are_font_objects(self):
        for f in [FONT_TITLE, FONT_SECTION, FONT_HEADER, FONT_DATA, FONT_KPI_BIG, FONT_KPI_MED, FONT_HINT]:
            self.assertIsInstance(f, Font)

class TestFillDefinitions(unittest.TestCase):
    def test_fills_are_pattern_fill_objects(self):
        for f in [FILL_HEADER, FILL_LIGHT, FILL_VERY_LIGHT, FILL_CARD]:
            self.assertIsInstance(f, PatternFill)

class TestApplyFunctions(unittest.TestCase):
    def setUp(self):
        self.wb = Workbook()
        self.ws = self.wb.active
    def test_apply_header_style(self):
        apply_header_style(self.ws, row=1, headers=["a","b"])
    def test_write_kpi_card(self):
        write_kpi_card(self.ws, row=1, col=1, label="L", value=100, unit="个")
        self.assertEqual(self.ws.cell(1,1).value, "L")
        self.assertEqual(self.ws.cell(1,2).value, 100)

if __name__ == "__main__":
    unittest.main()
