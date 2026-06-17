#!/usr/bin/env python3
"""DMS 非标询价周报自动化脚本（编排层 v3）。

架构总览（完整模式）：
  core/dms_browser.py      — Playwright 浏览器自动化（登录、筛选、提取）
  core/bom_parser.py       — BOM 物料解析（功率、容量计算）
  core/approval_parser.py  — 审批链信息解析
  core/orders_checker.py   — 下单检查
  core/excel_generator.py  — Excel 报表生成（4 Sheet）
  excel_styles.py          — Excel 样式主题
  column_definitions.py    — 列定义和常量集中管理
  generate_html_report.py  — HTML 报表生成
  dms_credentials.py       — 凭据管理

用法：
    python run_weekly_report.py [--headless] [--weeks N] [--workers N]
    python run_weekly_report.py --start-date 2026-05-01 --end-date 2026-05-31
    python run_weekly_report.py --stats-only --start-date 2026-06-01 --end-date 2026-06-07
"""

from __future__ import annotations

import argparse
import json
import subprocess
import asyncio
import glob
import logging
import os
import re
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path

from playwright.async_api import async_playwright
from playwright._impl._errors import TargetClosedError

# 共享模块
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import _compat  # noqa: F401 — side-effect: 修复 Windows 中文输出乱码
from column_definitions import (
    DMS_URL, NAV_TIMEOUT, LOAD_TIMEOUT, accumulate_power,
    STATUS_ORDERED, STATUS_NOT_ORDERED, STATUS_CHECK_FAILED,
    STATUS_YES, STATUS_NO, STATUS_NONE, STATUS_DASH, SHEET_DATA,
)
from core.dms_browser import (
    FlowRecord, TableProcessResult,
    do_login, filter_and_get_flow_ids, filter_and_get_flow_ids_via_api,
    extract_all_parallel, get_week_range,
    is_on_login_page,
)

logger = logging.getLogger("dms_report")
_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(logging.Formatter(
    "[%(asctime)s] %(levelname)s %(message)s", datefmt="%H:%M:%S"
))
if not logger.hasHandlers():
    logger.addHandler(_handler)

# ==================== 配置 ====================

USER_DATA_DIR = Path.home() / ".dms_browser_data"


def _find_headless_shell() -> str | None:
    """查找 Playwright headless shell 可执行文件路径。"""
    home = os.path.expanduser("~")
    patterns = [
        # Windows: chrome-win/headless_shell.exe
        os.path.join(home, "AppData", "Local", "ms-playwright",
                     "chromium_headless_shell-*", "chrome-win", "headless_shell.exe"),
        # Windows: chrome-headless-shell-win64/chrome-headless-shell.exe
        os.path.join(home, "AppData", "Local", "ms-playwright",
                     "chromium_headless_shell-*", "chrome-headless-shell-win64",
                     "chrome-headless-shell.exe"),
        # Linux
        os.path.join(home, ".cache", "ms-playwright",
                     "chromium_headless_shell-*", "chrome-linux", "headless_shell"),
        # macOS
        os.path.join(home, "Library", "Caches", "ms-playwright",
                     "chromium_headless_shell-*", "chrome-mac", "headless_shell"),
    ]
    for pattern in patterns:
        matches = glob.glob(pattern)
        if matches:
            return matches[0]
    return None


def configure_logging(verbose: bool = False) -> None:
    """根据 --verbose 标志配置日志级别。"""
    level = logging.DEBUG if verbose else logging.INFO
    logger.setLevel(level)
    _handler.setLevel(level)


# ==================== 终端摘要 ====================


def print_summary(
    start_time: datetime,
    start_date: str, end_date: str,
    flow_ids: list[str] | None = None,
    records: list[FlowRecord] | None = None,
    excel_path: str | None = None,
    error: str | None = None,
    discarded: int = 0,
    order_count: int = 0,
    order_date_range: str = "",
) -> None:
    """打印格式化执行摘要到终端。"""
    elapsed = (datetime.now() - start_time).total_seconds()
    ordered = sum(1 for r in (records or []) if r.ordered == STATUS_YES)
    not_ordered = sum(1 for r in (records or []) if r.ordered == STATUS_NO)
    check_failed = sum(1 for r in (records or []) if r.ordered == STATUS_CHECK_FAILED)

    print("\n========================================")
    print("  执行摘要")
    print("========================================")
    print(f"  查询范围    {start_date} ~ {end_date}")
    if flow_ids:
        print(f"  提取记录    {len(flow_ids)} 条")
    if discarded:
        print(f"  作废流程    {discarded} 条")
    if order_count:
        print(f"  订单总数    {order_count} 条（{order_date_range}）")
    if records:
        print(f"  已下单      {ordered} 条")
        print(f"  未下单      {not_ordered} 条")
        if check_failed:
            print(f"  检查失败    {check_failed} 条")
    if excel_path:
        print(f"  Excel文件   {excel_path}")
    if error:
        print(f"  执行状态    异常: {error}")
    print(f"  总耗时      {elapsed:.1f} 秒")
    print("========================================")


# ==================== 主流程编排 ====================


async def run(args: argparse.Namespace) -> None:
    """完整模式主流程：登录 → 筛选 → 提取 → 检查下单 → 生成报表。"""
    from core.orders_checker import check_orders_parallel
    from core.excel_generator import generate_excel

    output_dir = args.output_dir or os.getcwd()

    if args.start_date:
        start_date = args.start_date
        end_date = args.end_date or datetime.now().strftime("%Y-%m-%d")
    else:
        start_date, end_date = get_week_range(args.weeks)

    start_time = datetime.now()
    timestamp_str = start_time.strftime("%Y%m%d_%H%M%S")
    logger.info("=== 询价周报自动化（%d 并发）===", args.workers)
    logger.info("日期范围: %s ~ %s", start_date, end_date)
    logger.info("输出目录: %s", output_dir)
    os.makedirs(output_dir, exist_ok=True)

    flow_ids: list[str] = []
    records: list[FlowRecord] = []
    excel_path: str | None = None
    error_msg: str | None = None
    order_api_total: int = 0
    order_date_range: str = ""
    filter_result = None

    try:
        USER_DATA_DIR.mkdir(parents=True, exist_ok=True)
    except PermissionError:
        logger.error("无权限写入目录: %s，请检查权限或设置别的路径", USER_DATA_DIR)
        sys.exit(1)

    # 查找 headless shell 可执行文件
    headless_shell = _find_headless_shell()
    if not headless_shell:
        raise RuntimeError(
            "未找到 Playwright headless shell，请执行: playwright install chromium"
        )
    logger.debug("Headless shell: %s", headless_shell)

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            executable_path=headless_shell,
            user_data_dir=str(USER_DATA_DIR),
            headless=args.headless,
            no_viewport=True,
            locale="zh-CN",
            ignore_https_errors=True,
            args=["--start-maximized"],
        )
        page = await context.new_page()

        try:
            # 1. 登录
            await page.goto(DMS_URL, timeout=NAV_TIMEOUT)
            await page.wait_for_load_state("networkidle", timeout=LOAD_TIMEOUT)
            if is_on_login_page(page.url):
                await do_login(page)
            else:
                logger.info("会话有效（已复用缓存）")

            # 2. 筛选（优先 API 方式，失败时回退到 HTML 解析）
            filter_result = None
            try:
                logger.info("尝试 API 方式筛选流程列表...")
                filter_result = await filter_and_get_flow_ids_via_api(context, start_date, end_date)
                if filter_result and filter_result.flow_ids:
                    logger.info("API 筛选成功，获取 %d 个流程", len(filter_result.flow_ids))
                else:
                    logger.info("API 筛选返回空结果，回退到 HTML 解析")
                    filter_result = None
            except Exception as e:
                logger.warning("API 筛选失败（%s），回退到 HTML 解析", e)
                filter_result = None

            if not filter_result:
                filter_result = await filter_and_get_flow_ids(page, start_date, end_date)

            flow_ids = filter_result.flow_ids
            if not flow_ids:
                logger.info("本周无已办询价记录")
                return

            # 3. 并行提取详情 + 并行启动下单查询
            from core.orders_checker import fetch_ordered_flow_ids
            from column_definitions import ORDER_CHECK_EXTEND_DAYS

            all_details, (ordered_ids, order_api_total) = await asyncio.gather(
                extract_all_parallel(context, flow_ids, args.workers, flow_status=filter_result.flow_status),
                fetch_ordered_flow_ids(context, start_date, end_date),
            )
            if not all_details:
                logger.info("未能提取到任何详情")
                return

            # 4. 下单检查（ordered_ids 已在步骤 3 并发获取）
            for rec in all_details:
                rec.ordered = "是" if rec.flow_id in ordered_ids else "否"
            ordered_count = sum(1 for r in all_details if r.ordered == "是")
            logger.info("下单检查完成：%d 条已下单，%d 条未下单",
                        ordered_count, len(all_details) - ordered_count)
            records = all_details

            # 订单查询日期范围（含扩展天数）
            from datetime import datetime as _dt, timedelta as _td
            order_end_dt = _dt.strptime(end_date, "%Y-%m-%d") + _td(days=ORDER_CHECK_EXTEND_DAYS)
            order_date_range = f"{start_date} ~ {order_end_dt.strftime('%Y-%m-%d')}"

            # 5. 生成 Excel
            query_range = f"{start_date} ~ {end_date}"
            excel_path, rows_data = generate_excel(
                all_details, output_dir,
                query_range=query_range,
                timestamp_str=timestamp_str,
            )

            # 6. 生成 HTML 报表（不影响主流程）
            try:
                from generate_html_report import generate_html_report as _gen_html
                html_path = os.path.join(output_dir, f"询价周报报表_{timestamp_str}.html")
                _gen_html(rows_data, query_range, html_path)
                logger.info("HTML 报表已生成: %s", html_path)
            except Exception as html_e:
                logger.warning("HTML 报表生成失败（不影响 Excel）: %s", html_e)

        except KeyboardInterrupt:
            logger.info("用户中断执行")
            error_msg = "用户中断"
            discarded = filter_result.skipped_invalid if filter_result else 0
            print_summary(start_time, start_date, end_date, flow_ids, records, excel_path, error=error_msg, discarded=discarded, order_count=order_api_total, order_date_range=order_date_range)
            raise
        except Exception as e:
            error_msg = str(e)
            logger.error("执行异常: %s", e)
            traceback.print_exc()
        finally:
            try:
                await context.close()
            except Exception as e:
                logger.debug("浏览器上下文关闭时异常，忽略: %s", e)

    discarded = filter_result.skipped_invalid if filter_result else 0
    print_summary(start_time, start_date, end_date, flow_ids, records, excel_path, error=error_msg, discarded=discarded, order_count=order_api_total, order_date_range=order_date_range)


def stats_from_excel(args: argparse.Namespace) -> None:
    """仅统计模式主流程：读取已有 Excel → 按日期筛选 → 更新统计 Sheet。"""
    from core.excel_generator import (
        generate_excel, _update_summary_sheet, _fill_date_helper_column,
        _create_date_query_sheet_v2, _create_report_dashboard,
    )
    import openpyxl
    from excel_styles import (
        FONT_HEADER, FILL_HEADER, THIN_BORDER, ALIGN_HEADER,
    )
    from column_definitions import HEADERS

    output_dir = args.output_dir or os.getcwd()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 如果显式指定了输入文件，直接使用；否则自动查找
    if args.input_xlsx:
        file_path = os.path.abspath(args.input_xlsx)
    else:
        candidates = [
            os.path.join(output_dir, f"询价汇总_{ts}.xlsx"),
            os.path.join(output_dir, "询价汇总.xlsx"),
            os.path.join(output_dir, "询价汇总_v2.xlsx"),
        ]
        existing = [p for p in candidates if os.path.exists(p)]
        if len(existing) > 1:
            logger.warning("找到多个汇总文件，使用最新匹配: %s", existing[0])
        file_path = existing[0] if existing else candidates[0]

    if not os.path.exists(file_path):
        logger.error("未找到询价汇总文件: %s", file_path)
        return

    # 计算日期范围
    if args.this_month:
        today = datetime.now()
        start_date = today.replace(day=1).strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")
    elif args.start_date:
        start_date = args.start_date
        end_date = args.end_date or datetime.now().strftime("%Y-%m-%d")
    else:
        start_date, end_date = get_week_range(0)

    query_range = f"{start_date} ~ {end_date}"
    logger.info("=== 询价统计（仅统计模式）===")
    logger.info("统计范围: %s", query_range)
    logger.info("数据来源: %s", file_path)

    wb = openpyxl.load_workbook(file_path)
    data_ws = wb[SHEET_DATA]

    # 确保列头完整
    for col in range(1, len(HEADERS) + 1):
        existing = data_ws.cell(row=1, column=col).value
        header_value = HEADERS[col - 1]
        if existing is None or existing != header_value:
            cell = data_ws.cell(row=1, column=col, value=header_value)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_HEADER

    # 清除旧数据并补充辅助列
    for r in range(2, data_ws.max_row + 1):
        for c in range(15, 20):
            data_ws.cell(row=r, column=c).value = None
    _fill_date_helper_column(data_ws)

    # 按日期范围筛选
    filtered_rows: list = []
    for row in data_ws.iter_rows(min_row=2, values_only=True):
        flow_id = str(row[0]) if row[0] else ""
        if not re.match(r"^\d{15,}$", flow_id):
            continue
        submit_time = str(row[11]) if row[11] else ""
        if submit_time not in ("--", "无", ""):
            date_match = re.match(r"(\d{4}-\d{2}-\d{2})", submit_time)
            if date_match:
                row_date = date_match.group(1)
                if row_date < start_date or row_date > end_date:
                    continue
        filtered_rows.append(row)

    # 计算统计
    ordered_count = 0
    not_ordered_count = 0
    salesperson_set: set[str] = set()

    for row in filtered_rows:
        ordered = str(row[13] if row[13] else "")
        if ordered == STATUS_YES:
            ordered_count += 1
        else:
            not_ordered_count += 1
        sp = str(row[5] if row[5] else "")
        if sp not in (STATUS_DASH, STATUS_NONE, ""):
            salesperson_set.add(sp)

    total_module, total_inverter, total_battery = accumulate_power(filtered_rows)

    # 更新 Sheet
    _update_summary_sheet(wb, data_ws, query_range, filtered_rows=filtered_rows)
    _fill_date_helper_column(data_ws)
    _create_date_query_sheet_v2(wb)
    _create_report_dashboard(wb)

    save_path = os.path.join(output_dir, f"询价汇总_{ts}.xlsx")
    wb.save(save_path)

    # 终端输出
    print(f"\n{'=' * 45}")
    print(f"  📊 统计结果 ({query_range})")
    print(f"{'=' * 45}")
    print(f"  询价项目       {len(filtered_rows)} 个")
    print(f"  涉及业务员     {len(salesperson_set)} 人")
    print(f"  已下单         {ordered_count} 个")
    print(f"  未下单         {not_ordered_count} 个")
    if total_module > 0:
        print(f"  组件总功率     {total_module:.2f} kW")
    if total_inverter > 0:
        print(f"  逆变器总功率   {total_inverter:.2f} kW")
    if total_battery > 0:
        print(f"  电池总容量     {total_battery:.2f} kWh")
    ratio_display = f"{total_module / total_inverter:.2f}" if total_inverter > 0 else "--"
    print(f"  容配比         {ratio_display}")
    print("")
    logger.info("统计 Sheet 已更新，保存至: %s", save_path)


# ==================== CLI ====================


def main() -> None:
    parser = argparse.ArgumentParser(
        description="DMS 非标询价周报自动化 — 从 DMS 流程中心筛选询价、提取详情、检查下单、生成 Excel 报表。",
        epilog=(
            "使用示例:\n"
            "  %(prog)s --date-label \"本月\" --headless          # 本月数据，无头模式\n"
            "  %(prog)s --date-label \"上个月到现在\"              # 上个月至今\n"
            "  %(prog)s --start-date 2026-06-01 --end-date 2026-06-07  # 自定义日期\n"
            "  %(prog)s --weeks 1                                # 上周数据\n"
            "  %(prog)s --stats-only                             # 仅统计（跳过浏览器）\n"
            "  %(prog)s --stats-only --this-month                 # 仅统计本月\n"
            "\n"
            "日期标签（--date-label）支持: 本周 / 上周 / 本月 / 上月 / 本季度 / 去年 /\n"
            "  6月1号到6月7号 / 上个月12号到现在 / 上个月到现在 / 本季度 / 上季度 / 今年 / 去年 等自然语言格式。\n"
            "  优先级: --start-date > --date-label > --weeks > 默认本周"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--headless", action="store_true",
                        help="无头模式（不显示浏览器窗口），服务器环境或后台运行必选")
    parser.add_argument("--weeks", type=int, default=0,
                        help="查询最近 N 周（0=本周, 1=上周, 2=上上周…），默认 0")
    parser.add_argument("--start-date", type=str, default=None,
                        help="开始日期，格式 YYYY-MM-DD（例: 2026-06-01）。"
                             "与 --end-date 配合使用，设此值后 --weeks 失效")
    parser.add_argument("--end-date", type=str, default=None,
                        help="结束日期，格式 YYYY-MM-DD（例: 2026-06-07）。"
                             "不传则默认为今天")
    parser.add_argument("--date-label", type=str, default=None,
                        help="中文日期标签，自动解析。"
                             "支持: 本周/上周/本月/上月/本季度/上季度/今年/去年/上个月X号到现在/X月X号到X月X号等。"
                             "优先级低于 --start-date/--end-date，高于 --weeks")
    parser.add_argument("--workers", type=int, default=6,
                        help="并行提取并发数（1-8），默认 6。根据网络和 DMS 响应速度调整，"
                             "过高可能被限流")
    parser.add_argument("--output-dir", type=str, default=None,
                        help="输出目录（存放 Excel 和 HTML），默认当前工作目录")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="输出 DEBUG 级别详细日志，排查问题时使用")
    parser.add_argument("--stats-only", action="store_true",
                        help="仅统计模式：从已有 Excel 读取数据，按日期范围重新统计。"
                             "跳过浏览器登录和提取，快速出数")
    parser.add_argument("--input-xlsx", type=str, default=None,
                        help="仅统计模式下指定输入的 Excel 文件路径。"
                             "不传则自动查找 --output-dir 中的询价汇总文件")
    parser.add_argument("--this-month", action="store_true",
                        help="快捷统计本月（仅配合 --stats-only 使用），"
                             "等价于 --stats-only --date-label \"本月\"")
    args = parser.parse_args()

    configure_logging(args.verbose)

    # ───── 参数约束校验 ─────

    # workers 范围限制
    if args.workers < 1 or args.workers > 8:
        logger.error("--workers 并发数超出范围（允许 1-8），当前值: %d", args.workers)
        sys.exit(1)

    # 互斥检查: --stats-only + --this-month = --stats-only + --date-label 本月
    if args.stats_only and args.this_month:
        args.start_date, args.end_date = None, None
        args.date_label = "本月"

    # 三种日期方式只能选一种（优先级: start-date > date-label > weeks）
    if args.start_date and args.end_date:
        # 验证日期格式和逻辑
        try:
            sd = datetime.strptime(args.start_date, "%Y-%m-%d")
        except ValueError:
            logger.error("--start-date 格式无效，应为 YYYY-MM-DD，当前值: %s", args.start_date)
            sys.exit(1)
        try:
            ed = datetime.strptime(args.end_date, "%Y-%m-%d")
        except ValueError:
            logger.error("--end-date 格式无效，应为 YYYY-MM-DD，当前值: %s", args.end_date)
            sys.exit(1)
        if sd > ed:
            logger.error("--start-date (%s) 不能晚于 --end-date (%s)", args.start_date, args.end_date)
            sys.exit(1)
    elif args.date_label:
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            from _compat import captured_run
            result = captured_run(
                [sys.executable, os.path.join(script_dir, "resolve_date_range.py"),
                 args.date_label, "--json"],
                capture_output=True, text=True, check=True,
            )
            parsed = json.loads(result.stdout)
            args.start_date = parsed["start"]
            args.end_date = parsed["end"]
            logger.info("日期标签 '%s' → %s ~ %s", args.date_label, args.start_date, args.end_date)
        except Exception as e:
            logger.error("日期标签解析失败 '%s': %s", args.date_label, e)
            sys.exit(1)

    if args.stats_only:
        stats_from_excel(args)
    else:
        asyncio.run(run(args))


if __name__ == "__main__":
    main()
