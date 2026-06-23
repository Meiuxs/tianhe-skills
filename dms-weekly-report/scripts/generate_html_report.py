"""从 xlsx 数据生成 HTML 周报报表。

架构定位：
  本模块是报表渲染层，接收 Excel 原始数据行，输出独立 HTML 报表。
  数据来源：run_weekly_report.py 生成的「询价汇总」Sheet（19 列）。
  输出目标：references/report_template.html + 注入 JSON 数据 → 单文件 HTML。

数据管道：
  Excel 行 → read_rows_from_xlsx() → compute_rows_detail() → 有名字典列表 →
  注入模板 ROWS_DETAIL_JSON → 前端 JS 实时派生所有聚合统计。

用法：
    # 命令行
    python generate_html_report.py --xlsx 询价汇总.xlsx --output 报告.html --range "2026-06-01 ~ 2026-06-07"

    # 作为模块（被 run_weekly_report.py 调用）
    from generate_html_report import generate_html_report
    generate_html_report(rows_data, "2026-06-01 ~ 2026-06-07", "report.html")
"""

from __future__ import annotations

import html
import json
import math
import os
import re
import sys
from datetime import datetime
from typing import Any, TypedDict

# 导入共享的列定义
from column_definitions import (
    COL_FLOW_ID, COL_PROJECT_NAME, COL_PROVINCE, COL_SALESPERSON,
    COL_MODULE_KW, COL_INVERTER_KW, COL_BATTERY_KWH,
    COL_SUBMIT_TIME,
    COL_IS_VALID, COL_NEGOTIATION_PROCESSOR, COL_NEGOTIATION_STATUS, COL_NEGOTIATION_TIME,
    COL_PROVINCE_PROCESSOR, COL_PROVINCE_STATUS,
    COL_FINAL_APPROVAL_TIME,
    FLOW_ID_PATTERN,
)
class RowDetail(TypedDict):
    """单条询价数据行的类型定义，对应前端 ROWS_DETAIL 数组元素。

    注意：实际输出使用 camelCase 键（与前端 JS 保持一致），
    此处 TypedDict 也使用 camelCase 以匹配实际数据。
    """
    flowId: str
    projectName: str
    province: str
    salesperson: str
    modulePower: float
    inverterPower: float
    batteryCapacity: float
    isValid: str  # 新增：是否有效
    isInvalid: bool  # 新增：是否为作废流程
    negotiationApprover: str  # 新增：项目管理部核价审批人
    negotiationStatus: str  # 新增：项目管理部核价审批状态
    negotiationTime: str  # 新增：项目管理部核价审批时间
    submitDate: str
    finalDate: str
    provinceApprover: str
    provinceStatus: str



# ==================== 工具函数 ====================



def _format_datetime(value: Any) -> str:
    """将单元格日期值格式化为 YYYY-MM-DD HH:MM:SS 字符串。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    s = str(value)
    # 尝试截取常见日期时间格式 (YYYY-MM-DD HH:MM:SS 或 YYYY-MM-DDTHH:MM:SS)
    if len(s) >= 19 and s[4] == "-" and s[7] == "-":
        return s[:19].replace("T", " ")
    return s

def _safe_float(value: Any) -> float:
    """安全地将单元格值转为 float，处理字符串 "无" 等非数字值。"""
    if isinstance(value, (int, float)):
        result = float(value)
        if math.isinf(result) or math.isnan(result):
            return 0.0
        return result
    if isinstance(value, str):
        try:
            result = float(value)
            if math.isinf(result) or math.isnan(result):
                return 0.0
            return result
        except ValueError:
            return 0.0
    return 0.0


def _simple_replace(template: str, replacements: dict[str, str]) -> str:
    """将模板中的 {{KEY}} 替换为对应的 value。"""
    result = template
    for key, value in replacements.items():
        result = result.replace("{{" + key + "}}", str(value))
    return result


def _replace_json_field(template: str, field_name: str, data: Any) -> str:
    """将模板中的 {{FIELD_NAME_JSON}} 替换为 JSON 字符串。"""
    json_str = json.dumps(data, ensure_ascii=False, separators=(",", ": "))
    json_str = json_str.replace("<", "\\u003c").replace(">", "\\u003e").replace("/", "\\u002f")
    return template.replace("{{" + field_name + "_JSON}}", json_str)


# ==================== 数据读取 ====================


def read_rows_from_xlsx(xlsx_path: str) -> list[list[Any]]:
    """从 xlsx 的「询价汇总」Sheet 读取数据行（跳过表头）。"""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    try:
        if "询价汇总" not in wb.sheetnames:
            avail = ", ".join(wb.sheetnames)
            raise ValueError(
                f"工作簿中找不到「询价汇总」Sheet。可用 Sheet：{avail}"
            )
        ws = wb["询价汇总"]
        rows: list[list[Any]] = [
            [cell.value for cell in row]
            for row in ws.iter_rows(min_row=2)
        ]
    finally:
        wb.close()
    return rows


# ==================== 数据映射（唯一列索引引用点）====================


def compute_rows_detail(rows: list[list[Any]]) -> list[RowDetail]:
    """将原始数据行转为有名字典列表，供前端 ROWS_DETAIL 使用。

    所有列索引仅在此函数中出现一次，新增字段只需在此添加。
    """
    detail: list[RowDetail] = []
    for row in rows:
        raw_fid = row[COL_FLOW_ID]
        if raw_fid is None:
            continue
        if isinstance(raw_fid, float):
            # Excel 以 IEEE 754 双精度（约 15-16 位有效数字）存储数值型单元格。
            # 当流程编号 > 2^53（~9e15）时，最后几位不再精确。
            # 例如：12345678901234567890（20位）→ int(float) → 12345678901234567168。
            # 修复方法：在 Excel 中将该列格式设为「文本」后重新导出 xlsx。
            try:
                fid = str(int(raw_fid))
            except (ValueError, OverflowError):
                continue
        else:
            fid = str(raw_fid)
        if not re.match(FLOW_ID_PATTERN, fid):
            continue
        submit_time = _format_datetime(row[COL_SUBMIT_TIME])
        final_raw = _format_datetime(row[COL_FINAL_APPROVAL_TIME])
        
        # 获取是否有效字段
        is_valid = str(row[COL_IS_VALID]) if len(row) > COL_IS_VALID and row[COL_IS_VALID] else "否"
        # 获取项目管理部核价审批字段
        negotiation_approver = (
            str(row[COL_NEGOTIATION_PROCESSOR])
            if len(row) > COL_NEGOTIATION_PROCESSOR
            and row[COL_NEGOTIATION_PROCESSOR]
            and row[COL_NEGOTIATION_PROCESSOR] != "--"
            else ""
        )
        negotiation_status = (
            str(row[COL_NEGOTIATION_STATUS])
            if len(row) > COL_NEGOTIATION_STATUS
            and row[COL_NEGOTIATION_STATUS]
            and row[COL_NEGOTIATION_STATUS] != "--"
            else ""
        )
        negotiation_time = (
            str(row[COL_NEGOTIATION_TIME])
            if len(row) > COL_NEGOTIATION_TIME
            and row[COL_NEGOTIATION_TIME]
            and row[COL_NEGOTIATION_TIME] != "--"
            else ""
        )
        
        # 判断是否为作废流程（通过审批状态判断）
        is_invalid = "作废" in str(row[COL_PROVINCE_STATUS] or "") or "作废" in str(row[COL_NEGOTIATION_STATUS] or "")
        
        detail.append({
            "flowId": fid,
            "projectName": str(row[COL_PROJECT_NAME]) if row[COL_PROJECT_NAME] else "",
            "province": str(row[COL_PROVINCE]) if row[COL_PROVINCE] else "",
            "salesperson": str(row[COL_SALESPERSON]) if row[COL_SALESPERSON] else "",
            "modulePower": _safe_float(row[COL_MODULE_KW]),
            "inverterPower": _safe_float(row[COL_INVERTER_KW]),
            "batteryCapacity": _safe_float(row[COL_BATTERY_KWH]),
            "isValid": is_valid,
            "isInvalid": is_invalid,
            "negotiationApprover": negotiation_approver,
            "negotiationStatus": negotiation_status,
            "negotiationTime": negotiation_time,
            "submitDate": submit_time[:10] if len(submit_time) >= 10 else submit_time,
            "finalDate": final_raw[:10] if len(final_raw) >= 10 and final_raw not in ("--", "无", "") else "",
            "provinceApprover": (
                str(row[COL_PROVINCE_PROCESSOR])
                if row[COL_PROVINCE_PROCESSOR]
                and row[COL_PROVINCE_PROCESSOR] != "--"
                else ""
            ),
            "provinceStatus": (
                str(row[COL_PROVINCE_STATUS])
                if row[COL_PROVINCE_STATUS]
                and row[COL_PROVINCE_STATUS] != "--"
                else ""
            ),
        })
    return detail


# ==================== 主函数 ====================


def generate_html_report(
    rows_data: list[list[Any]],
    query_range: str,
    output_path: str,
    template_path: str | None = None,
) -> str:
    """从 rows_data 生成 HTML 报表文件，返回输出路径。

    Args:
        rows_data: 询价数据行，每行 19 列。
        query_range: 查询范围文本（如 "2026-06-01 ~ 2026-06-07"）。
        output_path: 输出 HTML 文件路径。
        template_path: 模板文件路径，默认使用 references/report_template.html。

    Returns:
        输出文件的路径。
    """
    if template_path is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(script_dir, "..", "references", "report_template.html")

    # 读取模板
    with open(template_path, "r", encoding="utf-8") as f:
        template = f.read()

    # 数据映射：Excel 原始行 → 有名字典列表
    rows_detail = compute_rows_detail(rows_data)

    # 空数据场景防御：无有效行时输出最小 HTML
    if not rows_detail:
        minimal_html = """<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><title>询价周报报表</title></head>
<body style="font-family: sans-serif; padding: 48px; text-align: center; color: #666;">
<h2>询价周报报表</h2>
<p>暂无数据</p>
<p style="font-size: 0.9em; color: #999;">数据范围：""" + html.escape(query_range) + """</p>
</body>
</html>"""
        output_dir = os.path.dirname(os.path.abspath(output_path))
        os.makedirs(output_dir, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(minimal_html)
        return output_path

    # 检查：是否有 float 类型的流程编号（可能因 Excel 数字格式丢失精度）
    # 注意：如果通过 SKILL.md 工作流运行，发现此提醒时应检查 xlsx 中流程编号列的单元格格式
    # ——须设为「文本」格式后重新导出，避免数字类型导致精度丢失。
    _float_fids = sum(
        1 for row in rows_data
        if isinstance(row[COL_FLOW_ID], float)
    )
    if _float_fids > 0:
        print(
            f"[提醒] 有 {_float_fids} 行的流程编号以数字格式存储在 Excel 中。"
            "超过 15 位的编号可能已丢失精度（建议在 Excel 中将该列设为「文本」后重新导出）。",
            file=sys.stderr,
        )

    now = datetime.now()
    now_str = now.strftime("%Y-%m-%d %H:%M")

    # 从 query_range 解析起止日期（格式 "2026-06-08 ~ 2026-06-12"）
    _date_match = re.match(r'(\d{4}-\d{2}-\d{2})\s*~\s*(\d{4}-\d{2}-\d{2})', query_range)
    query_start_date = _date_match.group(1) if _date_match else ''
    query_end_date = _date_match.group(2) if _date_match else ''

    # 替换纯文本标记（仅展示性字段，不涉及统计数据）
    replacements = {
        "REPORT_DATE_RANGE": query_range,
        "REPORT_GENERATED_AT": f"生成于 {now_str}",
        "DATA_SCOPE_TEXT": f"数据范围：{query_range} | 统计截止：{now.strftime('%Y-%m-%d')}",
        "FOOTER_TEXT": "询价周报报表 · 数据来源：DMS 流程中心 · 仅供内部参考",
        "SERVER_TIMESTAMP": now.isoformat(),
        "QUERY_START_DATE": query_start_date,
        "QUERY_END_DATE": query_end_date,
    }
    rendered = _simple_replace(template, replacements)

    # 注入 ROWS_DETAIL（唯一数据源，前端实时派生所有聚合）
    rendered = _replace_json_field(rendered, "ROWS_DETAIL", rows_detail)

    # 输出
    output_dir = os.path.dirname(os.path.abspath(output_path))
    os.makedirs(output_dir, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(rendered)

    return output_path


# ==================== CLI 入口 ====================


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="从 xlsx 生成 HTML 周报报表")
    parser.add_argument("--xlsx", required=True, help="输入的 xlsx 文件路径")
    parser.add_argument("--output", default="", help="输出的 html 文件路径（默认自动带时间戳）")
    parser.add_argument("--range", default="", help="查询范围文本，如 '2026-06-01 ~ 2026-06-07'")
    args = parser.parse_args()

    rows = read_rows_from_xlsx(args.xlsx)
    now = datetime.now()
    query_range = args.range or f"{now.strftime('%Y-%m-%d')} 数据"
    if not args.output:
        xlsx_dir = os.path.dirname(os.path.abspath(args.xlsx))
        args.output = f"{xlsx_dir}/询价周报报表_{now.strftime('%Y%m%d_%H%M%S')}.html"
    output = generate_html_report(rows, query_range, args.output)
    print(f"HTML 报表已生成：{output}")


if __name__ == "__main__":
    main()
