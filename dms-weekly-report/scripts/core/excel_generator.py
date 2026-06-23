"""Excel 报表生成模块。

生成包含以下 4 个 Sheet 的 Excel 报表：
  - 询价汇总：原始数据表
  - 询价统计：KPI 仪表盘式汇总
  - 日期查询：下拉交互式分时段统计
  - 数据看板：审批人统计/省公司排名/审批天数
"""

from __future__ import annotations

import logging
import os
import re
from datetime import date as dt_date, datetime, timedelta
from typing import Any

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

from column_definitions import (
    HEADERS, accumulate_power,
    COL_FLOW_ID, COL_PROJECT_NAME, COL_AGENT_CODE, COL_AGENT_NAME,
    COL_PROVINCE, COL_SALESPERSON,
    COL_MODULE_KW, COL_INVERTER_KW, COL_BATTERY_KWH,
    COL_UNIT_PRICE, COL_TOTAL_PRICE,
    COL_SUBMIT_TIME, COL_REMARK,
    COL_IS_VALID,
    COL_PROVINCE_PROCESSOR, COL_PROVINCE_STATUS,
    COL_NEGOTIATION_PROCESSOR, COL_NEGOTIATION_STATUS, COL_NEGOTIATION_TIME,
    COL_FINAL_APPROVAL_TIME,
    COL_FLOW_STATUS,
)
from excel_styles import COLUMN_WIDTHS
from column_definitions import FLOW_ID_PATTERN  # noqa: E402 — 流程编号正则常量
from excel_styles import (
    Colors,
    THIN_BORDER, BOTTOM_BORDER, CARD_BORDER,
    FONT_TITLE, FONT_SECTION, FONT_SUBSECTION, FONT_HEADER,
    FONT_DATA, FONT_LABEL, FONT_KPI_BIG, FONT_KPI_MED,
    FONT_HINT, FONT_VALUE, FONT_GREEN, FONT_ORANGE,
    FILL_HEADER, FILL_LIGHT, FILL_VERY_LIGHT, FILL_CARD, FILL_WHITE,
    ALIGN_CENTER, ALIGN_LEFT, ALIGN_RIGHT, ALIGN_HEADER, ALIGN_DATA,
    ROW_HEIGHT_TITLE, ROW_HEIGHT_SECTION, ROW_HEIGHT_DATA, ROW_HEIGHT_HEADER,
    apply_header_style, apply_data_row, apply_accent_row, apply_status_cell,
    write_section_title, write_kpi_card,
)

EXCEL_SERIAL_OFFSET = 693594  # Excel epoch (1899-12-30) to Gregorian ordinal baseline

logger = logging.getLogger("dms_report")


def generate_excel(
    records: list[Any],
    output_dir: str | None = None,
    query_range: str = "",
    timestamp_str: str | None = None,
) -> tuple[str, list[list[Any]]]:
    """生成格式化 Excel 文件（包含 4 个 Sheet）。

    Args:
        records: FlowRecord 对象列表。
        output_dir: 输出目录，默认当前工作目录。
        query_range: 查询范围字符串，如 "2026-06-01 ~ 2026-06-07"。
        timestamp_str: 时间戳字符串，默认自动生成。

    Returns:
        (文件路径, 行数据列表)
    """
    output_dir = output_dir or os.getcwd()
    ts = timestamp_str or datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(output_dir, f"询价汇总_{ts}.xlsx")
    backup_path = os.path.join(output_dir, f"询价汇总_{ts}_v2.xlsx")

    # 检查文件是否已存在（非本次运行创建），且可读写
    file_already_exists = os.path.exists(file_path)
    if file_already_exists:
        try:
            with open(file_path, "rb") as f:
                f.read(1)
        except (PermissionError, OSError):
            logger.warning("%s 被占用，使用备用文件名", file_path)
            file_path = backup_path
            file_already_exists = False

    # 构建行数据（19 列，含审批链信息）
    rows_data = _build_rows_data(records)

    if file_already_exists and os.path.exists(file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            rows_data = _deduplicate_rows(wb, rows_data)
            next_row = ws.max_row + 1
        except Exception as e:
            logger.warning("Excel 文件读取失败 (%s)，将重新创建: %s", file_path, e)
            wb = openpyxl.Workbook()
            ws = wb.active
            _init_worksheet(ws)
            next_row = 2
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        _init_worksheet(ws)
        next_row = 2

    # 长文本列（项目名称、代理商名称）左对齐，其余居中
    _TEXT_LEFT_COLS = (COL_PROJECT_NAME, COL_AGENT_NAME)

    for i, row_data in enumerate(rows_data):
        apply_data_row(ws, next_row, row_data, is_alt=(i % 2 == 1),
                       left_align_cols=_TEXT_LEFT_COLS)
        ws.row_dimensions[next_row].height = ROW_HEIGHT_DATA
        next_row += 1

    # 更新增强 Sheet
    _update_summary_sheet(wb, ws, query_range)
    _fill_date_helper_column(ws)
    _create_date_query_sheet_v2(wb)
    _create_report_dashboard(wb)

    try:
        wb.save(file_path)
    except PermissionError:
        backup_path = os.path.join(output_dir, f"询价汇总_{ts}_backup.xlsx")
        logger.warning("无法保存到 %s（文件被占用），尝试备用路径: %s", file_path, backup_path)
        wb.save(backup_path)
        file_path = backup_path
    except OSError as e:
        logger.error("保存 Excel 失败: %s", e)
        raise

    logger.info("Excel 已保存: %s (共 %d 条记录)", file_path, len(rows_data))
    return file_path, rows_data


# ==================== 辅助函数 ====================


def _init_worksheet(ws: Any) -> None:
    """初始化询价汇总 Sheet（表头、列宽、行高）。"""
    ws.title = "询价汇总"
    apply_header_style(ws, 1, HEADERS)
    for i, w in enumerate(COLUMN_WIDTHS):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w
    ws.row_dimensions[1].height = ROW_HEIGHT_HEADER


def _build_rows_data(records: list) -> list[list[Any]]:
    """从 FlowRecord 列表构建 21 列行数据。"""
    return [
        [
            r.flow_id, r.project_name,
            r.agent_code, r.agent_name,
            r.province, r.salesperson,
            r.module_kw, r.inverter_kw, r.battery_kwh,
            r.unit_price, r.total_price,
            r.submit_time, r.remark,
            r.is_valid,
            r.province_processor, r.province_status,
            r.negotiation_processor, r.negotiation_status, r.negotiation_time,
            r.final_approval_time,
            r.flow_status,
        ]
        for r in records
    ]


def _deduplicate_rows(wb: openpyxl.Workbook, rows_data: list[list[Any]]) -> list[list[Any]]:
    """去重：读取已有流程编号，过滤掉重复行。"""
    ws = wb.active
    existing_ids: set[str] = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and re.match(FLOW_ID_PATTERN, str(row[0])):
            existing_ids.add(str(row[0]))

    new_rows = [r for r in rows_data if str(r[0]) not in existing_ids]
    skipped = len(rows_data) - len(new_rows)
    if skipped:
        logger.info("跳过 %d 条重复记录", skipped)
    return new_rows


def _read_data_rows(data_ws: Any, max_cols: int = 21) -> list[list[Any]]:
    """从询价汇总 Sheet 读取数据行（跳过表头），供多个 Sheet 生成函数复用。

    Args:
        data_ws: 询价汇总 worksheet。
        max_cols: 读取的列数（默认 19 列，即 A-S）。

    Returns:
        二维列表，每行为一条记录。
    """
    last_data_row = max(data_ws.max_row, 2)
    rows_data: list[list[Any]] = []
    for r in range(2, last_data_row + 1):
        row = [data_ws.cell(r, c).value for c in range(1, max_cols + 1)]
        rows_data.append(row)
    return rows_data


def _fill_date_helper_column(ws: Any) -> None:
    """补充询价汇总Sheet的辅助列（日期序列号，供Excel公式使用），然后隐藏该列。

    注意：辅助列位于最后一列数据之后，避免覆盖现有数据。
    """
    from column_definitions import COL_FLOW_STATUS
    helper_col = COL_FLOW_STATUS + 2  # 在 flow_status 列之后留一列空白，使用下一列作为辅助列
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        existing = ws.cell(row=r, column=helper_col).value
        if existing is not None and isinstance(existing, (int, float)) and existing < 100000:
            continue
        l_val = ws.cell(row=r, column=12).value
        if l_val:
            date_match = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(l_val))
            if date_match:
                y, m, d = int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3))
                excel_serial = dt_date(y, m, d).toordinal() - EXCEL_SERIAL_OFFSET
                ws.cell(row=r, column=helper_col, value=excel_serial)
    # 隐藏辅助列 V
    ws.column_dimensions["V"].hidden = True


# ==================== 统计 Sheet ====================


def _update_summary_sheet(
    wb: Any, data_ws: Any, query_range: str,
    filtered_rows: list[Any] | None = None,
) -> None:
    """更新「询价统计」Sheet — 仪表盘式 KPI 布局。"""
    if "询价统计" in wb.sheetnames:
        del wb["询价统计"]

    ws = wb.create_sheet("询价统计")

    # 列宽：A=标签, B=数值, C=单位（紧凑布局，去掉间隔列）
    for c, w in enumerate([28, 20, 14], 1):
        ws.column_dimensions[chr(64 + c)].width = w

    # 计算统计数据
    total_projects = 0
    valid_count = 0
    invalid_count = 0
    salesperson_set: set[str] = set()
    valid_rows: list = []

    source_rows = filtered_rows if filtered_rows is not None else data_ws.iter_rows(min_row=2, values_only=True)

    for row in source_rows:
        flow_id = str(row[COL_FLOW_ID]) if row[COL_FLOW_ID] else ""
        if not re.match(r"^\d{15,}$", flow_id):
            continue
        total_projects += 1
        valid_rows.append(row)
        is_valid = str(row[COL_IS_VALID] if row[COL_IS_VALID] else "")
        if is_valid == "是":
            valid_count += 1
        else:
            invalid_count += 1
        sp = str(row[COL_SALESPERSON] if row[COL_SALESPERSON] else "")
        if sp not in ("--", "无", ""):
            salesperson_set.add(sp)

    total_module, total_inverter, total_battery = accumulate_power(valid_rows)

    # ---- 新排版：KPI 仪表盘 ----
    r = 1

    # 标题行
    ws.cell(r, 1, "询价统计汇总").font = FONT_TITLE
    ws.cell(r, 1).alignment = ALIGN_CENTER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = ROW_HEIGHT_TITLE
    r += 1

    # 统计周期
    ws.cell(r, 1, f"统计周期: {query_range}").font = FONT_VALUE
    ws.cell(r, 1).fill = FILL_LIGHT
    ws.cell(r, 1).alignment = ALIGN_CENTER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = ROW_HEIGHT_SECTION
    r += 2

    # ---- 区域1：询价概览 —— 每个 KPI 一行，从 A 列开始 ----
    r = write_section_title(ws, r, 1, "询价概览", merge_end_col=3)

    kpis_1 = [
        ("询价项目总数", f"{total_projects}", "个", FONT_KPI_BIG),
        ("涉及业务员", f"{len(salesperson_set)}", "人" if salesperson_set else "", FONT_KPI_MED),
        ("有效询价", f"{valid_count}", "个", FONT_GREEN),
        ("无效询价", f"{invalid_count}", "个", FONT_ORANGE),
    ]
    for i, (label, value, unit, vf) in enumerate(kpis_1):
        write_kpi_card(ws, r + i, 1, label, value, unit, vf)

    r += len(kpis_1) + 2

    # ---- 区域2：功率容量统计 —— 每个 KPI 一行，从 A 列开始 ----
    r = write_section_title(ws, r, 1, "功率容量统计", merge_end_col=3)

    module_display = f"{total_module:,.2f}" if total_module > 0 else "0"
    inverter_display = f"{total_inverter:,.2f}" if total_inverter > 0 else "0"
    battery_display = f"{total_battery:,.2f}" if total_battery > 0 else "0"
    ratio_display = f"{total_module / total_inverter:.2f}" if total_inverter > 0 else "--"

    kpis_2 = [
        ("组件总功率", module_display, "kW", FONT_KPI_BIG),
        ("逆变器总功率", inverter_display, "kW", FONT_KPI_MED),
        ("电池总容量", battery_display, "kWh", FONT_KPI_MED),
        ("容配比(组件/逆变器)", ratio_display, "", FONT_VALUE),
    ]
    for i, (label, value, unit, vf) in enumerate(kpis_2):
        write_kpi_card(ws, r + i, 1, label, value, unit, vf)

    r += len(kpis_2) + 2
    # 页脚说明
    ws.cell(r, 1, f"数据范围：全部历史数据 | 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}").font = FONT_HINT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)


# ==================== 日期查询 Sheet ====================


def _create_date_query_sheet_v2(wb: Any) -> None:
    """创建「日期查询」交互 Sheet — 紧凑布局，从 A 列开始。"""
    data_ws = wb["询价汇总"]
    rows_data = _read_data_rows(data_ws)

    def excel_serial(d: dt_date) -> int:
        return d.toordinal() - EXCEL_SERIAL_OFFSET

    def parse_date(l_val: Any) -> int | None:
        if l_val:
            m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(l_val))
            if m:
                return excel_serial(dt_date(int(m.group(1)), int(m.group(2)), int(m.group(3))))
        return None

    today = dt_date.today()
    wd = today.weekday()
    m_start = dt_date(today.year, today.month, 1)

    periods: dict[str, tuple[int, int]] = {}
    periods["全部"] = (excel_serial(dt_date(2000, 1, 1)), excel_serial(dt_date(2099, 12, 31)))
    periods["本周"] = (excel_serial(dt_date.fromordinal(today.toordinal() - wd)), excel_serial(today))
    periods["本月"] = (excel_serial(m_start), excel_serial(today))
    if today.month == 1:
        periods["上月"] = (excel_serial(dt_date(today.year - 1, 12, 1)), excel_serial(dt_date(today.year - 1, 12, 31)))
    else:
        lm = dt_date(today.year, today.month - 1, 1)
        lme = dt_date(today.year, today.month, 1) - dt_date.resolution
        periods["上月"] = (excel_serial(lm), excel_serial(lme))
    qs = (today.month - 1) // 3 * 3 + 1
    periods["本季度"] = (excel_serial(dt_date(today.year, qs, 1)), excel_serial(today))

    stats: dict[str, list[Any]] = {}
    for name, (s, e) in periods.items():
        cnt = 0
        mod = 0.0
        inv = 0.0
        bat = 0.0
        for row in rows_data:
            fid = str(row[COL_FLOW_ID]) if row[COL_FLOW_ID] else ""
            if not re.match(FLOW_ID_PATTERN, fid):
                continue
            # 按提交日期过滤当前时间段
            o = parse_date(row[COL_SUBMIT_TIME] if len(row) > COL_SUBMIT_TIME else None)
            if o is not None and (o < s or o > e):
                continue  # 有日期但不在当前时间段内，跳过
            if o is None and name != "全部":
                continue  # 无法解析日期的行仅纳入"全部"统计
            cnt += 1
            if len(row) > COL_MODULE_KW and isinstance(row[COL_MODULE_KW], (int, float)):
                mod += float(row[COL_MODULE_KW])
            if len(row) > COL_INVERTER_KW and isinstance(row[COL_INVERTER_KW], (int, float)):
                inv += float(row[COL_INVERTER_KW])
            if len(row) > COL_BATTERY_KWH and isinstance(row[COL_BATTERY_KWH], (int, float)):
                bat += float(row[COL_BATTERY_KWH])
        ratio = round(mod / inv, 2) if inv > 0 else 0
        stats[name] = [cnt, round(mod, 2), round(inv, 2), round(bat, 2), ratio]

    for sheet_name in ("日期查询", "日期查询(旧)"):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    ws = wb.create_sheet("日期查询")

    # 列宽 — 从 A 列开始紧凑布局
    for c, w in enumerate([24, 16, 20, 20, 20, 16, 18], 1):
        ws.column_dimensions[chr(64 + c)].width = w

    r = 1
    # 标题
    ws.cell(r, 1, "日期查询统计").font = FONT_TITLE
    ws.cell(r, 1).alignment = ALIGN_CENTER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = ROW_HEIGHT_TITLE
    r += 1

    # 筛选标签和下拉（从 B2 开始）
    for c in range(1, 4):
        ws.cell(r, c).fill = FILL_WHITE
        ws.cell(r, c).border = THIN_BORDER
    ws.cell(r, 1, "时间段筛选：").font = FONT_LABEL
    ws.cell(r, 1).alignment = ALIGN_LEFT
    ws.cell(r, 2, "全部").font = Font(bold=True, size=10, color=Colors.PRIMARY_BLUE, name="微软雅黑")
    ws.cell(r, 2).border = Border(
        left=Side("medium", color=Colors.PRIMARY_BLUE),
        right=Side("medium", color=Colors.PRIMARY_BLUE),
        top=Side("medium", color=Colors.PRIMARY_BLUE),
        bottom=Side("medium", color=Colors.PRIMARY_BLUE),
    )
    ws.cell(r, 2).alignment = ALIGN_CENTER
    ws.cell(r, 3, "  ← 点击选择").font = FONT_HINT
    ws.cell(r, 3).alignment = ALIGN_LEFT
    ws.row_dimensions[r].height = ROW_HEIGHT_DATA

    # 下拉数据验证
    presets = list(stats.keys())
    dv = openpyxl.worksheet.datavalidation.DataValidation(
        type="list", formula1=f'"{",".join(presets)}"',
        allow_blank=True,
    )
    dv.prompt = "选择时间段"
    dv.promptTitle = "快速选择"
    ws.add_data_validation(dv)
    dv.add(ws.cell(r, 2))
    r += 2

    # 预计算结果表
    r = write_section_title(ws, r, 1, "预计算结果", merge_end_col=7)

    headers = ["时间段", "项目数", "组件功率(kW)", "逆变器功率(kW)", "电池容量(kWh)", "容配比"]
    apply_header_style(ws, r, headers)
    table_header_row = r
    r += 1

    for i, name in enumerate(presets):
        vals = [name] + stats[name]
        fmts = [None, '0', '#,##0.00', '#,##0.00', '#,##0.00', '#,##0.00']
        apply_data_row(ws, r, vals, is_alt=(i % 2 == 1))
        for ci, nf in enumerate(fmts):
            if nf:
                ws.cell(r, 1 + ci).number_format = nf
        ws.row_dimensions[r].height = ROW_HEIGHT_DATA
        r += 1

    r += 1

    # INDEX/MATCH 公式结果（实时对应下拉选择）
    r = write_section_title(ws, r, 1, "当前选择结果", merge_end_col=7)

    n_p = len(presets)
    data_first = table_header_row + 1
    data_last = table_header_row + n_p
    dr = f"$A${data_first}:$F${data_last}"
    lr = f"$A${data_first}:$A${data_last}"

    for label, ci, nf in [
        ("询价项目数", 2, '0'),
        ("组件总功率 (kW)", 3, '#,##0.00'),
        ("逆变器总功率 (kW)", 4, '#,##0.00'),
        ("电池总容量 (kWh)", 5, '#,##0.00'),
        ("容配比（组件/逆变器）", 6, '#,##0.00'),
    ]:
        ws.cell(r, 1, label).font = FONT_LABEL
        ws.cell(r, 1).border = THIN_BORDER
        ws.cell(r, 1).alignment = ALIGN_CENTER
        ws.cell(r, 2).value = f'=IFERROR(INDEX({dr},MATCH($B$2,{lr},0),{ci}),0)'
        ws.cell(r, 2).font = Font(bold=True, size=10, color=Colors.TEXT_PRIMARY, name="微软雅黑")
        ws.cell(r, 2).border = THIN_BORDER
        ws.cell(r, 2).alignment = ALIGN_CENTER
        ws.cell(r, 2).number_format = nf
        ws.row_dimensions[r].height = ROW_HEIGHT_DATA
        r += 1

    r += 2
    # 使用说明
    r = write_section_title(ws, r, 1, "使用说明", merge_end_col=7)
    for tip in [
        "1. 点击 B2 → 从下拉选择预设时间段",
        "2. 预计算表格展示所有周期数据",
        "3. '当前选择结果' 随下拉实时变化",
        "4. 如需最新数据，重新运行周报脚本",
    ]:
        ws.cell(r, 1, tip).font = FONT_HINT
        r += 1


# ==================== 数据看板 Sheet ====================


def _create_report_dashboard(wb: Any) -> None:
    """创建「数据看板」Sheet — 管理层报表风格。"""
    if "数据看板" in wb.sheetnames:
        del wb["数据看板"]

    ws = wb.create_sheet("数据看板")

    data_ws = wb["询价汇总"]
    rows_data = _read_data_rows(data_ws)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    r = 1
    # 标题
    ws.cell(r, 1, "询价数据看板").font = FONT_TITLE
    ws.cell(r, 1).alignment = ALIGN_CENTER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = ROW_HEIGHT_TITLE
    r += 2

    # ===== 区域1：项目管理部核价审批统计（一行三卡片） =====
    r = write_section_title(ws, r, 1, "项目管理部核价审批统计", merge_end_col=7)

    negotiation_approved = 0
    negotiation_total = 0
    for row in rows_data:
        fid = str(row[COL_FLOW_ID]) if row[COL_FLOW_ID] else ""
        if not re.match(FLOW_ID_PATTERN, fid):
            continue
        proc = str(row[COL_NEGOTIATION_PROCESSOR] if len(row) > COL_NEGOTIATION_PROCESSOR and row[COL_NEGOTIATION_PROCESSOR] else "")
        status_val = str(row[COL_NEGOTIATION_STATUS] if len(row) > COL_NEGOTIATION_STATUS and row[COL_NEGOTIATION_STATUS] else "")
        if proc and proc not in ("--", "无", ""):
            negotiation_total += 1
            if "审批通过" in status_val:
                negotiation_approved += 1

    rate = f"{negotiation_approved/negotiation_total*100:.0f}%" if negotiation_total > 0 else "--"

    cards = [
        ("审批通过", negotiation_approved, "次", FONT_GREEN),
        ("经手总次数", negotiation_total, "次", FONT_KPI_MED),
        ("通过率", rate, "", FONT_VALUE),
    ]
    for i, (label, value, unit, vf) in enumerate(cards):
        col = 1 + i * 2  # A(1), C(3), E(5) — 从 A 列开始
        write_kpi_card(ws, r, col, label, value, unit, vf, alt_fill=True)
    r += 2

    # ===== 区域2：省公司询价排名 =====
    r = write_section_title(ws, r, 1, "省公司询价排名", merge_end_col=7)

    province_stats: dict[str, dict[str, Any]] = {}
    for row in rows_data:
        fid = str(row[COL_FLOW_ID]) if row[COL_FLOW_ID] else ""
        if not re.match(FLOW_ID_PATTERN, fid):
            continue
        pv = str(row[COL_PROVINCE] if len(row) > COL_PROVINCE and row[COL_PROVINCE] else "")
        if pv in ("--", "无", ""):
            continue
        g = float(row[COL_MODULE_KW]) if len(row) > COL_MODULE_KW and isinstance(row[COL_MODULE_KW], (int, float)) else 0
        if pv not in province_stats:
            province_stats[pv] = {"cnt": 0, "module": 0.0}
        province_stats[pv]["cnt"] += 1
        province_stats[pv]["module"] += g

    sorted_prov = sorted(province_stats.items(), key=lambda x: -x[1]["cnt"])

    headers = ["排名", "省公司", "询价次数", "组件总功率(kW)"]
    apply_header_style(ws, r, headers)
    r += 1

    for rank_i, (rank, (pv, data)) in enumerate(zip(range(1, len(sorted_prov) + 1), sorted_prov)):
        row_vals = [rank, pv, data["cnt"], round(data["module"], 2)]
        apply_data_row(ws, r, row_vals, is_alt=(rank_i % 2 == 1))
        ws.cell(r, 4).number_format = '#,##0.00'
        ws.row_dimensions[r].height = ROW_HEIGHT_DATA
        r += 1

    r += 2

    # ===== 区域3：审批天数统计（一行四卡片） =====
    r = write_section_title(ws, r, 1, "询价到审批完成天数", merge_end_col=7)

    days_list: list[int] = []
    for row in rows_data:
        fid = str(row[COL_FLOW_ID]) if row[COL_FLOW_ID] else ""
        if not re.match(FLOW_ID_PATTERN, fid):
            continue
        submit_time = str(row[COL_SUBMIT_TIME] if len(row) > COL_SUBMIT_TIME and row[COL_SUBMIT_TIME] else "")
        final_time = str(row[COL_FINAL_APPROVAL_TIME] if len(row) > COL_FINAL_APPROVAL_TIME and row[COL_FINAL_APPROVAL_TIME] else "")
        if submit_time in ("--", "") or final_time in ("--", ""):
            continue
        sm = re.match(r"(\d{4}-\d{2}-\d{2})", submit_time)
        fm = re.match(r"(\d{4}-\d{2}-\d{2})", final_time)
        if sm and fm:
            try:
                sd = datetime.strptime(sm.group(1), "%Y-%m-%d")
                fd = datetime.strptime(fm.group(1), "%Y-%m-%d")
                delta = (fd - sd).days
                if delta >= 0:
                    days_list.append(delta)
            except ValueError:
                pass

    avg_days = round(sum(days_list) / len(days_list), 1) if days_list else 0
    total_with_both = len(days_list)

    day_cards = [
        ("平均天数", avg_days, "天", FONT_VALUE),
        ("最短天数", min(days_list) if days_list else 0, "天", FONT_GREEN),
        ("最长天数", max(days_list) if days_list else 0, "天", FONT_ORANGE),
        ("统计样本数", total_with_both, "条", FONT_KPI_MED),
    ]
    for i, (label, value, unit, vf) in enumerate(day_cards):
        col = 1 + i * 2  # A(1), C(3), E(5), G(7)
        write_kpi_card(ws, r, col, label, value, unit, vf, alt_fill=True)
    r += 2

    # 说明
    r = write_section_title(ws, r, 1, "说明", merge_end_col=7)
    for tip in [
        "1. 项目管理部核价审批统计基于核价审批节点数据",
        "2. 省公司排名按询价次数降序排列",
        "3. 审批天数 = 审批完成时间 - 发起人提交审核时间",
        "4. 如需最新数据，重新运行周报脚本即可",
    ]:
        ws.cell(r, 1, tip).font = FONT_HINT
        r += 1
