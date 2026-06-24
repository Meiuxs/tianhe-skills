"""数据读取模块 — 数据源抽象。

包含：
  - XlsxDataReader: 从 xlsx 文件读取数据行的类
  - read_rows_from_xlsx: 同名函数（向后兼容，委托给 XlsxDataReader）
"""

from __future__ import annotations

import os
from typing import Any


class XlsxDataReader:
    """从 xlsx 文件读取「询价汇总」Sheet 数据行的数据源。

    用法：
        reader = XlsxDataReader("询价汇总.xlsx")
        rows = reader.read()

    支持上下文管理器：
        with XlsxDataReader("询价汇总.xlsx") as reader:
            rows = reader.read()
    """

    def __init__(self, file_path: str):
        self.file_path = os.path.abspath(file_path)

    def __enter__(self) -> XlsxDataReader:
        return self

    def __exit__(self, *args: Any) -> None:
        pass

    def read(self) -> list[list[Any]]:
        """从 xlsx 的「询价汇总」Sheet 读取数据行（跳过表头）。"""
        import openpyxl

        wb = openpyxl.load_workbook(self.file_path)
        try:
            if "询价汇总" not in wb.sheetnames:
                avail = ", ".join(wb.sheetnames)
                raise ValueError(
                    f"工作簿中找不到「询价汇总」Sheet。可用 Sheet：{avail}"
                )
            ws = wb["询价汇总"]
            rows: list[list[Any]] = [
                [cell.value for cell in row]
                for row in ws.iter_rows(min_row=2)
            ]
        finally:
            wb.close()
        return rows


def read_rows_from_xlsx(xlsx_path: str) -> list[list[Any]]:
    """从 xlsx 读取数据行（委托给 XlsxDataReader，向后兼容）。"""
    return XlsxDataReader(xlsx_path).read()
